import json
from io import BytesIO

from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.utils import timezone
from django.db.models import Count, Sum, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from .models import SolicitudBaja, Producto, Categoria, Promocion
from apps.users.decorators import solo_vendedor, solo_administrador, rol_requerido

# ==================== SOLICITUDES DE BAJA POR VENDEDOR ====================

@solo_vendedor
@require_http_methods(["POST"])
def solicitar_baja_api(request):
    """
    POST /productos/api/solicitar-baja/
    
    Vendedor solicita una baja de producto por daño, expiración, etc.
    
    Body JSON:
    {
        "producto_id": 1,
        "cantidad": 5,
        "motivo": "Producto dañado/Expirado/Otro"
    }
    
    Respuesta:
    {
        "success": true,
        "baja_id": 10,
        "estado": "PENDIENTE",
        "mensaje": "Solicitud enviada al administrador"
    }
    
    Nota: El stock se descuenta solo cuando el admin aprueba.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            producto_id = data.get('producto_id')
            cantidad = int(data.get('cantidad', 0))
            motivo = data.get('motivo', '').strip()

            # Validaciones
            if not producto_id or cantidad <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'Producto e cantidad son requeridos'
                }, status=400)

            if not motivo:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe proporcionar un motivo'
                }, status=400)

            # Validar que el producto existe
            try:
                producto = Producto.objects.get(id=producto_id)
            except Producto.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Producto no encontrado'
                }, status=404)

            # Validar stock disponible
            if producto.stock_actual < cantidad:
                return JsonResponse({
                    'success': False,
                    'error': f'Stock insuficiente. Disponible: {producto.stock_actual}'
                }, status=400)

            # Crear solicitud (con transacción)
            with transaction.atomic():
                baja = SolicitudBaja.objects.create(
                    producto=producto,
                    cantidad=cantidad,
                    motivo=motivo,
                    solicitado_por=request.user,
                    estado='PENDIENTE'
                )

            return JsonResponse({
                'success': True,
                'baja_id': baja.id,
                'estado': baja.estado,
                'mensaje': f'✅ Solicitud de baja enviada. ID: {baja.id}'
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'JSON inválido'
            }, status=400)
        except Exception as e:
            print(f"Error en solicitar_baja_api: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    }, status=405)


# ==================== ADMINISTRACIÓN DE SOLICITUDES DE BAJA ====================

@solo_vendedor
def mis_solicitudes_baja(request):
    """
    GET /productos/mis-solicitudes-baja/
    
    Página para que el vendedor vea el estado de sus solicitudes de baja.
    Soporta filtros por estado y paginación de 10.
    """
    from django.core.paginator import Paginator
    
    filtro_estado = request.GET.get('estado', '')
    
    bajas = SolicitudBaja.objects.select_related(
        'producto', 'solicitado_por', 'revisado_por'
    ).filter(solicitado_por=request.user).order_by('-fecha_solicitud')
    
    if filtro_estado:
        bajas = bajas.filter(estado=filtro_estado)
    
    stats = {
        'pendientes': SolicitudBaja.objects.filter(solicitado_por=request.user, estado='PENDIENTE').count(),
        'aprobadas': SolicitudBaja.objects.filter(solicitado_por=request.user, estado='APROBADO').count(),
        'rechazadas': SolicitudBaja.objects.filter(solicitado_por=request.user, estado='RECHAZADO').count(),
    }
    
    paginator = Paginator(bajas, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'products/mis_solicitudes_baja.html', {
        'bajas': page_obj,
        'stats': stats,
        'filtro_activo': filtro_estado,
    })


@solo_administrador
def bajas_pendientes(request):
    """
    GET /productos/bajas/pendientes/
    
    Bandeja de entrada del admin: Solicitudes de baja pendientes.
    Muestra últimas 10 solicitudes con paginación.
    
    Acciones posibles:
    - Aprobar (descuenta stock)
    - Rechazar (no afecta stock)
    
    Información mostrada:
    - Producto, cantidad solicitada, motivo
    - Vendedor solicitante, fecha
    - Botones de acción
    """
    from django.core.paginator import Paginator
    
    # Opciones de filtrado - por defecto PENDIENTE
    filtro_estado = request.GET.get('estado', 'PENDIENTE')
    filtro_usuario = request.GET.get('usuario', '')
    
    # Query base
    bajas = SolicitudBaja.objects.select_related(
        'producto', 'solicitado_por', 'revisado_por'
    ).order_by('-fecha_solicitud')
    
    # Aplicar filtro de estado
    if filtro_estado:
        bajas = bajas.filter(estado=filtro_estado)
    
    # Aplicar filtro de usuario
    if filtro_usuario:
        bajas = bajas.filter(solicitado_por__username__icontains=filtro_usuario)
    
    # Paginación: 10 items por página
    paginator = Paginator(bajas, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'pendientes': SolicitudBaja.objects.filter(estado='PENDIENTE').count(),
        'aprobadas': SolicitudBaja.objects.filter(estado='APROBADO').count(),
        'rechazadas': SolicitudBaja.objects.filter(estado='RECHAZADO').count(),
    }
    
    # Lista de vendedores para el filtro
    vendedores = User.objects.filter(perfil__rol='vendedor').order_by('first_name', 'username')

    return render(request, 'products/bajas_pendientes.html', {
        'bajas': page_obj,
        'stats': stats,
        'filtro_activo': filtro_estado,
        'filtro_usuario': filtro_usuario,
        'vendedores': vendedores,
    })


@solo_administrador
@require_http_methods(["POST"])
def aprobar_baja(request, baja_id):
    """
    POST /productos/api/bajas/<id>/aprobar/
    
    Admin aprueba una solicitud de baja.
    
    Acciones:
    - Descuenta el stock del producto
    - Marca solicitud como APROBADO
    - Registra quién aprobó y cuándo
    
    Respuesta:
    {
        "success": true,
        "mensaje": "Baja aprobada. Stock actualizado.",
        "producto": "Nombre Producto",
        "cantidad": 5,
        "stock_nuevo": 95
    }
    """
    try:
        baja = SolicitudBaja.objects.select_related('producto').get(id=baja_id)

        # Validar que está pendiente
        if baja.estado != 'PENDIENTE':
            return JsonResponse({
                'success': False,
                'error': f'Esta solicitud ya fue {baja.estado.lower()}'
            }, status=400)

        # Transacción atómica: descuentostock + marca aprobada
        with transaction.atomic():
            # Validar stock una última vez (por si cambió)
            if baja.producto.stock_actual < baja.cantidad:
                return JsonResponse({
                    'success': False,
                    'error': f'Stock insuficiente. Solo hay {baja.producto.stock_actual}'
                }, status=400)

            # Descontar stock
            baja.producto._stock_motivo = 'BAJA'
            baja.producto._stock_usuario = request.user
            baja.producto._stock_referencia = f"baja:{baja.id}"
            baja.producto.stock_actual -= baja.cantidad
            baja.producto.save()

            # Marcar como aprobada
            baja.estado = 'APROBADO'
            baja.revisado_por = request.user
            baja.save()

        return JsonResponse({
            'success': True,
            'mensaje': f'✅ Baja aprobada. Stock actualizado.',
            'producto': baja.producto.nombre,
            'cantidad': baja.cantidad,
            'stock_nuevo': baja.producto.stock_actual
        })

    except SolicitudBaja.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Solicitud de baja no encontrada'
        }, status=404)
    except Exception as e:
        print(f"Error en aprobar_baja: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@solo_administrador
@require_http_methods(["POST"])
def rechazar_baja(request, baja_id):
    """
    POST /productos/api/bajas/<id>/rechazar/
    
    Admin rechaza una solicitud de baja.
    
    Body JSON:
    {
        "comentario": "Motivo del rechazo (opcional)"
    }
    
    Acciones:
    - Marca como RECHAZADO
    - Guarda comentario del admin
    - El stock NO se ve afectado
    
    Respuesta:
    {
        "success": true,
        "mensaje": "Baja rechazada",
        "producto": "Nombre Producto"
    }
    """
    try:
        data = json.loads(request.body) if request.body else {}
        comentario = data.get('comentario', '').strip()

        baja = SolicitudBaja.objects.select_related('producto').get(id=baja_id)

        # Validar que está pendiente
        if baja.estado != 'PENDIENTE':
            return JsonResponse({
                'success': False,
                'error': f'Esta solicitud ya fue {baja.estado.lower()}'
            }, status=400)

        # Marcar como rechazada
        baja.estado = 'RECHAZADO'
        baja.revisado_por = request.user
        baja.comentario_admin = comentario
        baja.save()

        return JsonResponse({
            'success': True,
            'mensaje': f'✅ Baja rechazada',
            'producto': baja.producto.nombre
        })

    except SolicitudBaja.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Solicitud de baja no encontrada'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'JSON inválido'
        }, status=400)
    except Exception as e:
        print(f"Error en rechazar_baja: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def _inventario_queryset(request):
    q = (request.GET.get('q') or '').strip()
    low = (request.GET.get('low') or '') in ('1', 'true', 'on')
    cat = (request.GET.get('cat') or '').strip()
    order = (request.GET.get('order') or 'nombre').strip()

    productos = Producto.objects.select_related('categoria').con_ventas()

    if q:
        productos = productos.filter(
            Q(nombre__icontains=q)
            | Q(code__icontains=q)
            | Q(categoria__nombre__icontains=q)
        )

    if cat:
        productos = productos.filter(categoria_id=cat)

    if low:
        productos = productos.stock_bajo()

    if order == 'mas_vendidos':
        productos = productos.mas_vendidos()
    elif order == 'stock':
        productos = productos.order_by('stock_actual', 'nombre')
    elif order == 'categoria':
        productos = productos.order_by('categoria__nombre', 'nombre')
    else:
        productos = productos.order_by('nombre')

    filtros = {
        'q': q,
        'low': low,
        'cat': cat,
        'order': order,
    }

    return productos, filtros


@rol_requerido('admin', 'vendedor')
@require_http_methods(["GET"])
def inventario(request):
    productos, filtros = _inventario_queryset(request)
    categorias = Categoria.objects.order_by('nombre')

    return render(request, 'products/inventario_list.html', {
        'productos': productos,
        'categorias': categorias,
        'filtros': filtros,
        'es_admin': getattr(request, 'es_admin', False),
        'es_vendedor': getattr(request, 'es_vendedor', False),
    })


@rol_requerido('admin', 'vendedor')
@require_http_methods(["GET"])
def inventario_table(request):
    productos, filtros = _inventario_queryset(request)
    return render(request, 'products/partials/inventario_table.html', {
        'productos': productos,
        'filtros': filtros,
        'es_admin': getattr(request, 'es_admin', False),
        'es_vendedor': getattr(request, 'es_vendedor', False),
    })


@rol_requerido('admin', 'vendedor')
@require_http_methods(["POST"])
def ajustar_stock(request, producto_id: int):
    """
    API para ajustar stock. 
    - Admin: puede sumar o restar (ajuste completo)
    - Vendedor: solo puede añadir stock (entradas), no restar
    """
    es_admin = getattr(request, 'es_admin', False)
    
    try:
        delta = int(request.POST.get('delta', '0'))
    except ValueError:
        return HttpResponse("Delta inválido", status=400)

    if delta == 0:
        return HttpResponse(status=204)
    
    if not es_admin and delta < 0:
        return HttpResponse("No tienes permiso para reducir stock. Contacta al administrador.", status=403)

    try:
        with transaction.atomic():
            producto = Producto.objects.select_for_update().select_related('categoria').get(id=producto_id)

            stock_nuevo = producto.stock_actual + delta
            if stock_nuevo < 0:
                return HttpResponse("Stock insuficiente", status=400)

            motivo = 'ENTRADA' if delta > 0 else 'AJUSTE'
            producto._stock_motivo = motivo
            producto._stock_usuario = request.user
            producto._stock_referencia = "inventario_ui"
            producto.stock_actual = stock_nuevo
            producto.save()

        producto = Producto.objects.select_related('categoria').con_ventas().get(id=producto_id)
        return render(request, 'products/partials/inventario_row.html', {
            'p': producto,
            'es_admin': es_admin,
            'es_vendedor': getattr(request, 'es_vendedor', False),
        })
    except Producto.DoesNotExist:
        return HttpResponse("Producto no encontrado", status=404)


@rol_requerido('admin', 'vendedor')
@require_http_methods(["GET", "POST"])
def crear_producto(request):
    if request.method == 'POST':
        from decimal import Decimal
        from django.core.exceptions import ValidationError
        
        nombre = request.POST.get('nombre', '').strip()
        code = request.POST.get('code', '').strip()
        categoria_id = request.POST.get('categoria')
        stock_actual = int(request.POST.get('stock_actual', 0) or 0)
        stock_minimo = int(request.POST.get('stock_minimo', 5) or 5)
        costo_compra = Decimal(request.POST.get('costo_compra', '0') or 0)
        precio_venta = Decimal(request.POST.get('precio_venta', '0') or 0)
        
        if not nombre or not code or not categoria_id:
            return render(request, 'products/producto_form.html', {
                'error': 'Todos los campos marcados con * son requeridos',
                'categorias': Categoria.objects.all(),
            }, status=400)
        
        try:
            categoria = Categoria.objects.get(id=categoria_id)
        except Categoria.DoesNotExist:
            return render(request, 'products/producto_form.html', {
                'error': 'Categoría no válida',
                'categorias': Categoria.objects.all(),
            }, status=400)
        
        if Producto.objects.filter(code=code).exists():
            return render(request, 'products/producto_form.html', {
                'error': 'Ya existe un producto con este código',
                'categorias': Categoria.objects.all(),
            }, status=400)
        
        producto = Producto.objects.create(
            nombre=nombre,
            code=code,
            categoria=categoria,
            stock_actual=stock_actual,
            stock_minimo=stock_minimo,
            costo_compra=costo_compra,
            precio_venta=precio_venta,
        )
        
        return redirect('products:inventario')
    
    categorias = Categoria.objects.all()
    return render(request, 'products/producto_form.html', {
        'categorias': categorias,
    })


@solo_administrador
@require_http_methods(["GET", "POST"])
def editar_producto(request, producto_id):
    try:
        producto = Producto.objects.get(id=producto_id)
    except Producto.DoesNotExist:
        return redirect('products:inventario')
    
    if request.method == 'POST':
        from decimal import Decimal
        
        nombre = request.POST.get('nombre', '').strip()
        code = request.POST.get('code', '').strip()
        categoria_id = request.POST.get('categoria')
        stock_actual = int(request.POST.get('stock_actual', 0) or 0)
        stock_minimo = int(request.POST.get('stock_minimo', 5) or 5)
        costo_compra = Decimal(request.POST.get('costo_compra', '0') or 0)
        precio_venta = Decimal(request.POST.get('precio_venta', '0') or 0)
        
        if not nombre or not code or not categoria_id:
            return render(request, 'products/producto_form.html', {
                'producto': producto,
                'error': 'Todos los campos marcados con * son requeridos',
                'categorias': Categoria.objects.all(),
            }, status=400)
        
        try:
            categoria = Categoria.objects.get(id=categoria_id)
        except Categoria.DoesNotExist:
            return render(request, 'products/producto_form.html', {
                'producto': producto,
                'error': 'Categoría no válida',
                'categorias': Categoria.objects.all(),
            }, status=400)
        
        if Producto.objects.filter(code=code).exclude(id=producto_id).exists():
            return render(request, 'products/producto_form.html', {
                'producto': producto,
                'error': 'Ya existe otro producto con este código',
                'categorias': Categoria.objects.all(),
            }, status=400)
        
        producto.nombre = nombre
        producto.code = code
        producto.categoria = categoria
        producto.stock_actual = stock_actual
        producto.stock_minimo = stock_minimo
        producto.costo_compra = costo_compra
        producto.precio_venta = precio_venta
        producto.save()
        
        return redirect('products:inventario')
    
    categorias = Categoria.objects.all()
    return render(request, 'products/producto_form.html', {
        'producto': producto,
        'categorias': categorias,
    })


@solo_administrador
@require_http_methods(["POST"])
def eliminar_producto(request, producto_id):
    try:
        producto = Producto.objects.get(id=producto_id)
        producto.delete()
        return JsonResponse({'success': True, 'mensaje': 'Producto eliminado correctamente'})
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==================== REPORTES (ADMIN) ====================

@solo_administrador
@require_http_methods(["GET"])
def inventario_pdf(request):
    from collections import defaultdict
    productos = Producto.objects.select_related('categoria').order_by('categoria__nombre', 'nombre')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Inventario")
    styles = getSampleStyleSheet()

    fecha = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    elementos = [
        Paragraph("📦 REPORTE GENERAL DE INVENTARIO", styles["Title"]),
        Paragraph(f"Generado: {fecha}", styles["Normal"]),
        Spacer(1, 12),
    ]
    
    # Agrupar productos por categoría
    productos_por_categoria = defaultdict(list)
    for p in productos:
        cat_nombre = getattr(p.categoria, 'nombre', 'Sin categoría')
        productos_por_categoria[cat_nombre].append(p)
    
    # Crear tabla por categoría
    grand_total_stock = 0
    grand_total_valor = 0
    
    for categoria, productos_list in productos_por_categoria.items():
        # Encabezado de categoría
        elementos.append(Paragraph(f"<b>📂 {categoria}</b>", styles["Heading2"]))
        elementos.append(Spacer(1, 6))
        
        # Cabecera de la tabla
        data = [["Código", "Producto", "Stock", "PVP", "Costo", "Subtotal"]]
        
        subtotal_stock = 0
        subtotal_valor = 0
        
        for p in productos_list:
            subtotal = p.stock_actual * p.precio_venta
            subtotal_stock += p.stock_actual
            subtotal_valor += float(subtotal)
            
            data.append([
                p.code,
                p.nombre[:30],
                str(p.stock_actual),
                f"${float(p.precio_venta):,.2f}",
                f"${float(p.costo_compra):,.2f}",
                f"${float(subtotal):,.2f}",
            ])
        
        # Fila de subtotal por categoría
        data.append(["", "SUBTOTAL", str(subtotal_stock), "", "", f"${subtotal_valor:,.2f}"])
        
        tabla = Table(data, colWidths=[0.9*inch, 2*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.9*inch])
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#334155")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.HexColor("#eef2ff")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef3c7")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))
        
        elementos.append(tabla)
        elementos.append(Spacer(1, 12))
        
        grand_total_stock += subtotal_stock
        grand_total_valor += subtotal_valor
    
    # Total general
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph("=" * 40, styles["Normal"]))
    elementos.append(Paragraph(f"<b>TOTAL GENERAL - Stock: {grand_total_stock} unidades - Valor: ${grand_total_valor:,.2f}</b>", styles["Heading2"]))
    
    doc.build(elementos)

    buffer.seek(0)
    filename = f"inventario_{timezone.localdate().isoformat()}.pdf"
    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@solo_administrador
@require_http_methods(["GET"])
def inventario_excel(request):
    from collections import defaultdict
    productos = Producto.objects.select_related('categoria').order_by('categoria__nombre', 'nombre')

    wb = Workbook()
    
    # Agrupar productos por categoría
    productos_por_categoria = defaultdict(list)
    for p in productos:
        cat_nombre = getattr(p.categoria, 'nombre', 'Sin categoría')
        productos_por_categoria[cat_nombre].append(p)
    
    grand_total_stock = 0
    grand_total_valor = 0
    
    for categoria, productos_list in productos_por_categoria.items():
        # Crear hoja por categoría (nombre limitado a 31 chars)
        ws = wb.create_sheet(title=categoria[:31] if len(categoria) > 31 else categoria)
        
        headers = ["Código", "Producto", "Stock", "PVP", "Costo", "Subtotal"]
        ws.append(headers)
        
        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        subtotal_stock = 0
        subtotal_valor = 0
        
        for p in productos_list:
            subtotal = p.stock_actual * p.precio_venta
            subtotal_stock += p.stock_actual
            subtotal_valor += float(subtotal)
            
            ws.append([
                p.code,
                p.nombre,
                int(p.stock_actual),
                float(p.precio_venta),
                float(p.costo_compra),
                float(subtotal),
            ])
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        grand_total_stock += subtotal_stock
        grand_total_valor += subtotal_valor
    
    # Hoja de resumen
    if len(wb.sheetnames) > 1:
        ws_resumen = wb.create_sheet(title="Resumen")
        ws_resumen.append(["REPORTE GENERAL DE INVENTARIO"])
        ws_resumen.append([f"Fecha: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}"])
        ws_resumen.append([])
        ws_resumen.append(["Categoría", "Stock", "Valor Total"])
        
        for cat in productos_por_categoria.keys():
            stock_cat = sum(p.stock_actual for p in productos_por_categoria[cat])
            valor_cat = sum(p.stock_actual * p.precio_venta for p in productos_por_categoria[cat])
            ws_resumen.append([cat, stock_cat, round(valor_cat, 2)])
        
        ws_resumen.append([])
        ws_resumen.append(["TOTAL GENERAL", grand_total_stock, round(grand_total_valor, 2)])
        
        # Aplicar formato a la hoja resumen
        for row in ws_resumen.iter_rows():
            for cell in row:
                if row[0].row == 1:
                    cell.font = Font(bold=True, size=14)
                elif row[0].row == 4:
                    cell.font = Font(bold=True)
    
    # Eliminar hoja por defecto si hay otras hojas
    if len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])
    
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"inventario_{timezone.localdate().isoformat()}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ==================== GESTIÓN DE PROMOCIONES ====================

@solo_administrador
def listar_promociones(request):
    """
    GET /productos/promociones/
    Lista todas las promociones.
    """
    from django.utils import timezone
    
    filtro = request.GET.get('filtro', '').strip()
    hoy = timezone.now().date()
    
    promociones = Promocion.objects.select_related('producto').order_by('-activa', '-fecha_inicio')
    
    if filtro == 'activas':
        promociones = promociones.filter(activa=True, fecha_inicio__lte=hoy, fecha_fin__gte=hoy)
    elif filtro == 'vencidas':
        promociones = promociones.filter(fecha_fin__lt=hoy)
    elif filtro == 'inactivas':
        promociones = promociones.filter(activa=False)
    
    activas_count = Promocion.objects.filter(activa=True, fecha_inicio__lte=hoy, fecha_fin__gte=hoy).count()
    vencidas_count = Promocion.objects.filter(fecha_fin__lt=hoy).count()
    inactivas_count = Promocion.objects.filter(activa=False).count()
    
    return render(request, 'products/promociones/list_promociones.html', {
        'promociones': promociones,
        'activas_count': activas_count,
        'vencidas_count': vencidas_count,
        'inactivas_count': inactivas_count,
        'filtro_actual': filtro,
        'hoy': hoy,
    })


@solo_administrador
def crear_promocion(request):
    """
    GET/POST /productos/promociones/crear/
    Crea una nueva promoción.
    """
    if request.method == 'POST':
        from decimal import Decimal
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        producto_id = request.POST.get('producto')
        tipo_descuento = request.POST.get('tipo_descuento')
        valor_descuento = Decimal(str(request.POST.get('valor_descuento')))
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        producto = Producto.objects.get(id=producto_id)
        
        if tipo_descuento == 'PORCENTAJE':
            precio_promocional = producto.precio_venta * (Decimal('1') - valor_descuento / Decimal('100'))
        else:
            precio_promocional = producto.precio_venta - valor_descuento
        
        promocion = Promocion.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            producto=producto,
            tipo_descuento=tipo_descuento,
            valor_descuento=valor_descuento,
            precio_promocional=precio_promocional,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            activa=True
        )
        
        return redirect('products:listar_promociones')
    
    productos = Producto.objects.all()
    return render(request, 'products/promociones/create_pro.html', {
        'productos': productos,
    })


@solo_administrador
@require_http_methods(["POST"])
def toggle_promocion(request, promocion_id):
    """
    POST /productos/promociones/<id>/toggle/
    Activa o desactiva una promoción.
    """
    promocion = Promocion.objects.get(id=promocion_id)
    promocion.activa = not promocion.activa
    promocion.save()
    
    return JsonResponse({'success': True, 'activa': promocion.activa})


@solo_administrador
def editar_promocion(request, promocion_id):
    """
    GET/POST /productos/promociones/<id>/editar/
    Edita una promoción existente.
    """
    try:
        promocion = Promocion.objects.select_related('producto').get(id=promocion_id)
    except Promocion.DoesNotExist:
        messages.error(request, 'Promoción no encontrada.')
        return redirect('products:listar_promociones')
    
    if request.method == 'POST':
        from decimal import Decimal
        
        promocion.nombre = request.POST.get('nombre')
        promocion.descripcion = request.POST.get('descripcion', '')
        promocion.producto_id = request.POST.get('producto')
        promocion.tipo_descuento = request.POST.get('tipo_descuento')
        promocion.valor_descuento = Decimal(str(request.POST.get('valor_descuento')))
        promocion.fecha_inicio = request.POST.get('fecha_inicio')
        promocion.fecha_fin = request.POST.get('fecha_fin')
        
        producto = Producto.objects.get(id=promocion.producto_id)
        
        if promocion.tipo_descuento == 'PORCENTAJE':
            promocion.precio_promocional = producto.precio_venta * (Decimal('1') - promocion.valor_descuento / Decimal('100'))
        elif promocion.tipo_descuento == 'FIJO':
            promocion.precio_promocional = producto.precio_venta - promocion.valor_descuento
        elif promocion.tipo_descuento == '2X1':
            promocion.precio_promocional = producto.precio_venta
        
        promocion.save()
        
        messages.success(request, f'Promoción "{promocion.nombre}" actualizada exitosamente.')
        return redirect('products:listar_promociones')
    
    productos = Producto.objects.all()
    return render(request, 'products/promociones/edit_pro.html', {
        'promocion': promocion,
        'productos': productos,
    })
